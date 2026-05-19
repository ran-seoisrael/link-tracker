import io
import logging
import os
from typing import Optional

import openpyxl
from fastapi import BackgroundTasks, FastAPI, File, Form, Request, UploadFile, status
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

import auth
import database as db
from checker import check_link
from scheduler import init_scheduler, shutdown_scheduler

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(title="Link Tracker")
templates = Jinja2Templates(directory="templates")
app.mount("/static", StaticFiles(directory="static"), name="static")


# ── Startup / Shutdown ────────────────────────────────────────────────────

@app.on_event("startup")
def startup():
    db.init_db()
    _seed_initial_user()
    init_scheduler()


@app.on_event("shutdown")
def shutdown():
    shutdown_scheduler()


def _seed_initial_user():
    """Create the default admin user if not exists."""
    email = os.environ.get("ADMIN_EMAIL", "rans@seoisrael.co.il")
    password = os.environ.get("ADMIN_PASSWORD", "Seoisrael2024!")
    name = os.environ.get("ADMIN_NAME", "רן")
    if not db.get_user_by_email(email):
        db.create_user(email, name, auth.hash_password(password))
        logger.info("Created initial user: %s", email)


# ── Auth helpers ──────────────────────────────────────────────────────────

def get_current_user(request: Request):
    token = request.cookies.get(auth.COOKIE_NAME)
    if not token:
        return None
    user_id = auth.decode_token(token)
    if not user_id:
        return None
    row = db.get_user_by_id(user_id)
    return dict(row) if row else None


def require_user(request: Request):
    user = get_current_user(request)
    if not user:
        return None
    return user


# ── Login / Logout ────────────────────────────────────────────────────────

@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request, error: str = ""):
    return templates.TemplateResponse("login.html", {"request": request, "error": error})


@app.post("/login")
def login(request: Request, email: str = Form(...), password: str = Form(...)):
    user = db.get_user_by_email(email)
    if not user or not auth.verify_password(password, user["password_hash"]):
        return templates.TemplateResponse(
            "login.html",
            {"request": request, "error": "אימייל או סיסמה שגויים"},
            status_code=400,
        )
    token = auth.create_token(user["id"])
    response = RedirectResponse("/", status_code=303)
    response.set_cookie(
        auth.COOKIE_NAME, token,
        httponly=True, samesite="lax",
        max_age=60 * 60 * 24 * 30,
    )
    return response


@app.get("/logout")
def logout():
    response = RedirectResponse("/login", status_code=303)
    response.delete_cookie(auth.COOKIE_NAME)
    return response


# ── Dashboard ─────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
def dashboard(request: Request, filter_user: Optional[int] = None,
              filter_status: Optional[str] = None):
    user = require_user(request)
    if not user:
        return RedirectResponse("/login", status_code=303)

    links = db.list_links(filter_user)
    users = db.list_users()

    # client-side status filter applied in template via JS
    return templates.TemplateResponse("dashboard.html", {
        "request": request,
        "user": user,
        "links": [dict(l) for l in links],
        "users": [dict(u) for u in users],
        "filter_user": filter_user,
        "filter_status": filter_status,
    })


# ── Add link ──────────────────────────────────────────────────────────────

@app.get("/add", response_class=HTMLResponse)
def add_page(request: Request):
    user = require_user(request)
    if not user:
        return RedirectResponse("/login", status_code=303)
    users = db.list_users()
    return templates.TemplateResponse("add_edit.html", {
        "request": request, "user": user,
        "users": [dict(u) for u in users],
        "link": None, "error": "",
    })


@app.post("/add")
def add_link(
    request: Request,
    page_url: str = Form(...),
    expected_link_url: str = Form(...),
    expected_anchor: str = Form(...),
    notes: str = Form(""),
    owner_user_id: Optional[int] = Form(None),
):
    user = require_user(request)
    if not user:
        return RedirectResponse("/login", status_code=303)
    target_user = owner_user_id if owner_user_id else user["id"]
    db.create_link(target_user, page_url, expected_link_url, expected_anchor, notes)
    return RedirectResponse("/", status_code=303)


# ── Edit link ─────────────────────────────────────────────────────────────

@app.get("/edit/{link_id}", response_class=HTMLResponse)
def edit_page(request: Request, link_id: int):
    user = require_user(request)
    if not user:
        return RedirectResponse("/login", status_code=303)
    link = db.get_link(link_id)
    if not link:
        return RedirectResponse("/", status_code=303)
    users = db.list_users()
    return templates.TemplateResponse("add_edit.html", {
        "request": request, "user": user,
        "users": [dict(u) for u in users],
        "link": dict(link), "error": "",
    })


@app.post("/edit/{link_id}")
def edit_link(
    request: Request, link_id: int,
    page_url: str = Form(...),
    expected_link_url: str = Form(...),
    expected_anchor: str = Form(...),
    notes: str = Form(""),
):
    user = require_user(request)
    if not user:
        return RedirectResponse("/login", status_code=303)
    db.update_link(link_id, page_url, expected_link_url, expected_anchor, notes, user["id"])
    return RedirectResponse("/", status_code=303)


# ── Delete ────────────────────────────────────────────────────────────────

@app.post("/delete/{link_id}")
def delete_link(request: Request, link_id: int):
    user = require_user(request)
    if not user:
        return RedirectResponse("/login", status_code=303)
    db.delete_link(link_id, user["id"])
    return RedirectResponse("/", status_code=303)


# ── Check single link ─────────────────────────────────────────────────────

@app.post("/check/{link_id}")
def check_single(request: Request, link_id: int, background_tasks: BackgroundTasks):
    user = require_user(request)
    if not user:
        return RedirectResponse("/login", status_code=303)

    def _do_check():
        row = db.get_link(link_id)
        if not row:
            return
        result = check_link(row["page_url"], row["expected_link_url"], row["expected_anchor"])
        db.update_link_check(link_id, result)

    background_tasks.add_task(_do_check)
    return RedirectResponse(f"/?checking={link_id}", status_code=303)


# ── Check all ─────────────────────────────────────────────────────────────

@app.post("/check-all")
def check_all(request: Request, background_tasks: BackgroundTasks):
    user = require_user(request)
    if not user:
        return RedirectResponse("/login", status_code=303)

    def _do_all():
        ids = db.get_all_link_ids()
        for link_id in ids:
            row = db.get_link(link_id)
            if not row:
                continue
            try:
                result = check_link(row["page_url"], row["expected_link_url"], row["expected_anchor"])
                db.update_link_check(link_id, result)
            except Exception as exc:
                logger.error("check_all error on link %d: %s", link_id, exc)

    background_tasks.add_task(_do_all)
    return RedirectResponse("/?checking=all", status_code=303)


# ── Export ────────────────────────────────────────────────────────────────

EXPORT_HEADERS = [
    "ID", "עובד", "כתובת עמוד", "אנקור טקסט", "כתובת קישור",
    "סטטוס 200", "סריקה", "אינדוקס", "קנוניקל", "אנקור נמצא",
    "URL תואם", "שגיאות", "נבדק לאחרונה",
]


def _link_to_row(l: dict) -> list:
    def b(v):
        if v is None:
            return "—"
        return "✓" if v else "✗"

    return [
        l["id"], l.get("user_name", ""), l["page_url"], l["expected_anchor"],
        l["expected_link_url"],
        b(l["check_status_200"]), b(l["check_crawlable"]),
        b(l["check_indexable"]), b(l["check_canonical"]),
        b(l["check_anchor_found"]), b(l["check_url_match"]),
        l.get("check_errors") or "", l.get("last_checked") or "",
    ]


@app.get("/export/csv")
def export_csv(request: Request):
    user = require_user(request)
    if not user:
        return RedirectResponse("/login", status_code=303)
    links = [dict(l) for l in db.list_links()]
    output = io.StringIO()
    # BOM for Excel Hebrew support
    output.write("﻿")
    import csv
    writer = csv.writer(output)
    writer.writerow(EXPORT_HEADERS)
    for l in links:
        writer.writerow(_link_to_row(l))
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=links.csv"},
    )


@app.get("/export/excel")
def export_excel(request: Request):
    user = require_user(request)
    if not user:
        return RedirectResponse("/login", status_code=303)
    links = [dict(l) for l in db.list_links()]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "קישורים"
    ws.sheet_view.rightToLeft = True
    ws.append(EXPORT_HEADERS)
    for l in links:
        ws.append(_link_to_row(l))
    # Style header row
    from openpyxl.styles import Font, PatternFill
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=links.xlsx"},
    )


# ── Change password ───────────────────────────────────────────────────────

@app.get("/change-password", response_class=HTMLResponse)
def change_password_page(request: Request):
    user = require_user(request)
    if not user:
        return RedirectResponse("/login", status_code=303)
    return templates.TemplateResponse("change_password.html", {
        "request": request, "user": user, "error": "", "success": False,
    })


@app.post("/change-password")
def change_password(
    request: Request,
    current_password: str = Form(...),
    new_password: str = Form(...),
    confirm_password: str = Form(...),
):
    user = require_user(request)
    if not user:
        return RedirectResponse("/login", status_code=303)
    ctx = {"request": request, "user": user, "success": False}
    if not auth.verify_password(current_password, user["password_hash"]):
        return templates.TemplateResponse("change_password.html", {**ctx, "error": "הסיסמה הנוכחית שגויה"})
    if new_password != confirm_password:
        return templates.TemplateResponse("change_password.html", {**ctx, "error": "הסיסמאות החדשות אינן תואמות"})
    if len(new_password) < 6:
        return templates.TemplateResponse("change_password.html", {**ctx, "error": "הסיסמה חייבת להכיל לפחות 6 תווים"})
    with db.get_db() as conn:
        conn.execute(
            "UPDATE users SET password_hash = ? WHERE id = ?",
            (auth.hash_password(new_password), user["id"]),
        )
    return templates.TemplateResponse("change_password.html", {**ctx, "error": "", "success": True})


# ── Import from Excel ─────────────────────────────────────────────────────

@app.get("/import", response_class=HTMLResponse)
def import_page(request: Request):
    user = require_user(request)
    if not user:
        return RedirectResponse("/login", status_code=303)
    return templates.TemplateResponse("import.html", {
        "request": request, "user": user,
        "result": None, "error": "",
    })


@app.post("/import")
async def import_excel(
    request: Request,
    file: UploadFile = File(...),
    skip_empty: bool = Form(True),
):
    user = require_user(request)
    if not user:
        return RedirectResponse("/login", status_code=303)

    ctx = {"request": request, "user": user}

    if not file.filename.endswith((".xlsx", ".xls")):
        return templates.TemplateResponse("import.html", {
            **ctx, "result": None,
            "error": "יש להעלות קובץ Excel (.xlsx בלבד)",
        })

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
    except Exception as exc:
        return templates.TemplateResponse("import.html", {
            **ctx, "result": None,
            "error": f"שגיאה בפתיחת הקובץ: {exc}",
        })

    imported = 0
    skipped = 0
    errors = []

    # col index (0-based): B=1, C=2, D=3, E=4, F=5, G=6
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        def cell(idx):
            try:
                v = row[idx]
                return str(v).strip() if v is not None else ""
            except IndexError:
                return ""

        employee_name = cell(1)   # B
        client_name   = cell(2)   # C
        client_domain = cell(3)   # D
        page_url      = cell(4)   # E
        link_url      = cell(5)   # F
        anchor        = cell(6)   # G

        # skip rows missing the three required fields
        if not page_url or not link_url or not anchor:
            skipped += 1
            continue

        # basic URL validation
        if not page_url.startswith("http"):
            errors.append(f"שורה {i}: כתובת עמוד לא תקינה — {page_url[:60]}")
            skipped += 1
            continue

        notes_parts = []
        if client_name:
            notes_parts.append(f"לקוח: {client_name}")
        if client_domain:
            notes_parts.append(f"דומיין: {client_domain}")
        if employee_name:
            notes_parts.append(f"עובד מקורי: {employee_name}")
        notes = " | ".join(notes_parts)

        try:
            db.create_link(user["id"], page_url, link_url, anchor, notes)
            imported += 1
        except Exception as exc:
            errors.append(f"שורה {i}: {exc}")
            skipped += 1

    return templates.TemplateResponse("import.html", {
        **ctx,
        "result": {"imported": imported, "skipped": skipped, "errors": errors},
        "error": "",
    })
